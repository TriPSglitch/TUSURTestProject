from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from excel.Styles import *
import os

from Cities import *

filepath = os.path.join('test.xlsx')


def get_recommendations(*, values_list: list) -> str:
    recommendations = ''

    if 'дождь' in values_list[3]:
        recommendations += 'возьмите зонт|'

    if float(values_list[10]) > 3.0:
        recommendations += '|наденьте головной убор|'

    if float(values_list[8]) <= 2.0:
        recommendations += '|будьте внимательнее на дороге'

    if '||' in recommendations:
        recommendations = recommendations.replace('||', ', ')

    if '|' in recommendations:
        recommendations = recommendations.replace('|', '')

    if recommendations == '':
        recommendations = '\t'

    return recommendations


def get_headers(*, dict_list: dict) -> list:
    header = list()
    header.append('Дата и время')

    for value in dict_list.values():
        for head in value:
            for key in head.keys():
                if key not in header:
                    header.append(key)

    return header


def get_values(*, key: str, value: list, header: list) -> list:
    items = list()
    items.append(key)

    for row in value:
        for col in header:
            if col in row:
                items.append(row[col])

    if len(items) > 2:
        items.append(get_recommendations(values_list=items))

    return items


def fill_excel(*, dict_list: dict, city: City, ws: Worksheet, wb: Workbook):
    city_name = list()
    city_name.append(city.name)
    ws.append(city_name)

    last_row = ws.max_row
    max_column = 0
    for value in dict_list.values():
        for row in value:
            if len(row) > 1:
                max_column = len(row) + 2
            else:
                max_column = len(row) + 1

    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row,
                   end_column=max_column)
    for cell in ws[last_row]:
        cell.border = get_border_style()

    ws.cell(row=last_row, column=1).style = get_city_style()

    if dict_list:
        header = get_headers(dict_list=dict_list)
        if max_column > 2:
            header.append('Рекомендации')
        ws.append(header)

        for key, value in dict_list.items():
            items = get_values(key=key, value=value, header=header)
            ws.append(items)

    ws.append(list())
    ws.append(list())

    for cell in ws[last_row + 1]:
        cell.style = get_header_style()

    for row in ws.iter_rows(min_row=last_row + 2, max_row=ws.max_row, min_col=1,
                            max_col=max_column):
        for cell in row:
            cell.style = get_cell_style()

    for col in ws.columns:
        column = col[1].column_letter
        ws.column_dimensions[column].width = get_columns_width(columns=col)

    wb.save(filepath)
    return filepath
